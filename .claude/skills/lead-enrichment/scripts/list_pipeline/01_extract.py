# -*- coding: utf-8 -*-
"""Stage 1 — Extract: xlsx/csv → normalized extracted.json.

Usage:
    python 01_extract.py --input my_list.xlsx --output extracted.json
    python 01_extract.py --input my_list.xlsx --sheets "Лист1,Лист2"
    python 01_extract.py --input my_list.csv --output extracted.json
"""
import sys, io, os, re, json, csv, argparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Header aliases: any of these (case-insensitive, trim, ignore non-word) → canonical key
HEADER_ALIASES = {
    'name':              ['name', 'company', 'org', 'орг', 'компания', 'наименование', 'название'],
    'inn':               ['inn', 'инн', 'taxid', 'tax_id'],
    'ul':                ['ul', 'юл', 'юрлицо', 'юр.лицо', 'юр_лицо', 'legalname', 'legal_name', 'полноенаименование'],
    'phone':             ['phone', 'телефон', 'тел', 'phoneorg', 'phone_org', 'phone1', 'phone_org1', 'orgphone'],
    'email':             ['email', 'mail', 'почта', 'e-mail'],
    'site':              ['site', 'website', 'сайт', 'url', 'web'],
    'rev2024':           ['rev2024', 'revenue2024', 'выручка2024', 'выручка_2024', 'revenue', 'выручка'],
    'rev2023':           ['rev2023', 'revenue2023', 'выручка2023', 'выручка_2023'],
    'industry':          ['industry', 'отрасль', 'okved', 'оквэд'],
    'segment':           ['segment', 'сегмент', 'category', 'категория'],
    'last_call_comment': ['comment', 'комментарий', 'last_call_comment', 'комментарийзвонка', 'примечание', 'note'],
}


def normalize_header(h):
    if h is None:
        return ''
    s = str(h).lower().strip()
    s = s.replace('ё', 'е')
    s = re.sub(r'[\s\-\.\(\)]+', '', s)
    return s


def map_headers(row_headers):
    """Return mapping {col_index: canonical_key} for matched columns."""
    mapping = {}
    for idx, h in enumerate(row_headers):
        nh = normalize_header(h)
        if not nh:
            continue
        for canonical, aliases in HEADER_ALIASES.items():
            if nh in aliases or any(a in nh for a in aliases if len(a) >= 4):
                mapping[idx] = canonical
                break
    return mapping


INN_RE = re.compile(r'\D')


def normalize_inn(v):
    if v is None:
        return ''
    s = str(v).strip()
    digits = INN_RE.sub('', s)
    if len(digits) in (10, 12):
        return digits
    return ''


REV_RE = re.compile(r'[^\d,.\-]')


def normalize_revenue(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower()
    multiplier = 1.0
    if 'млрд' in s or 'bn' in s:
        multiplier = 1_000_000_000.0
    elif 'млн' in s or 'mn' in s:
        multiplier = 1_000_000.0
    elif 'тыс' in s or 'k' == s[-1:]:
        multiplier = 1_000.0
    s = REV_RE.sub('', s).replace(',', '.')
    # If multiple dots — keep only the last as decimal separator
    if s.count('.') > 1:
        parts = s.split('.')
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        return float(s) * multiplier
    except ValueError:
        return None


def extract_xlsx(path, sheets_filter):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = []
    sheets = wb.sheetnames if not sheets_filter else [s for s in wb.sheetnames if s in sheets_filter]
    for sheet_name in sheets:
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        headers = sheet_rows[0]
        mapping = map_headers(headers)
        if 'inn' not in mapping.values() and 'name' not in mapping.values():
            print(f'  [WARN] sheet "{sheet_name}": no inn/name columns recognized, skipped', file=sys.stderr)
            continue
        for r_idx, row in enumerate(sheet_rows[1:], start=2):
            rec = {'source': sheet_name, 'row_idx': r_idx}
            for col_idx, val in enumerate(row):
                key = mapping.get(col_idx)
                if not key or val is None:
                    continue
                if key == 'inn':
                    rec['inn'] = normalize_inn(val)
                elif key in ('rev2024', 'rev2023'):
                    rev = normalize_revenue(val)
                    if rev is not None:
                        rec[key] = rev
                else:
                    s = str(val).strip()
                    if s:
                        rec[key] = s
            if rec.get('inn') or rec.get('name'):
                rows.append(rec)
    return rows


def extract_csv(path):
    rows = []
    with open(path, encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f)
        all_rows = list(reader)
    if not all_rows:
        return rows
    headers = all_rows[0]
    mapping = map_headers(headers)
    if 'inn' not in mapping.values() and 'name' not in mapping.values():
        print(f'  [WARN] csv: no inn/name columns recognized', file=sys.stderr)
        return rows
    sheet_name = os.path.splitext(os.path.basename(path))[0]
    for r_idx, row in enumerate(all_rows[1:], start=2):
        rec = {'source': sheet_name, 'row_idx': r_idx}
        for col_idx, val in enumerate(row):
            key = mapping.get(col_idx)
            if not key or val is None or val == '':
                continue
            if key == 'inn':
                rec['inn'] = normalize_inn(val)
            elif key in ('rev2024', 'rev2023'):
                rev = normalize_revenue(val)
                if rev is not None:
                    rec[key] = rev
            else:
                s = str(val).strip()
                if s:
                    rec[key] = s
        if rec.get('inn') or rec.get('name'):
            rows.append(rec)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True, help='xlsx or csv file')
    ap.add_argument('--output', default='extracted.json', help='output JSON path')
    ap.add_argument('--sheets', default='', help='comma-separated sheet names (xlsx only); empty=all')
    args = ap.parse_args()

    input_path = os.path.abspath(args.input)
    if not os.path.exists(input_path):
        print(f'ERROR: {input_path} not found', file=sys.stderr)
        sys.exit(1)

    sheets_filter = [s.strip() for s in args.sheets.split(',') if s.strip()]
    ext = os.path.splitext(input_path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        rows = extract_xlsx(input_path, sheets_filter)
    elif ext == '.csv':
        rows = extract_csv(input_path)
    else:
        print(f'ERROR: unsupported extension {ext}. Use .xlsx or .csv', file=sys.stderr)
        sys.exit(1)

    out_path = os.path.abspath(args.output)
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    inn_count = sum(1 for r in rows if r.get('inn'))
    sources = sorted({r['source'] for r in rows})
    print(f'Extracted: {len(rows)} rows ({inn_count} with INN) from {len(sources)} source(s): {", ".join(sources)}')
    print(f'  → {out_path}')


if __name__ == '__main__':
    main()
