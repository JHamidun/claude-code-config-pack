# -*- coding: utf-8 -*-
"""Stage 4 — Build: assemble final enriched.xlsx with all 10-12 sheets.

Usage:
    python 04_build_xlsx.py \
        --input extracted.json \
        --bitrix bitrix_data.json \
        --output enriched.xlsx \
        [--source-xlsx original.xlsx]   # copy original sheets and augment them
        [--research research/results.json]
        [--research-queue research/queue.json]
        [--linkedin linkedin.json]
        [--bitrix-url https://we.company.example]

linkedin.json format:
    [{"name": "Сидоров А.А.", "role": "CIO", "company": "ООО Альфа",
      "linkedin": "https://...", "reaction": "Принял инвайт", "segment": "Финансы",
      "screen": "...", "notes": "..."}]
"""
import sys, io, os, json, re, shutil, argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')


# ========== Styles ==========
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
MATCH_FILL = PatternFill('solid', fgColor='E2EFDA')   # Tier B
HOT_FILL   = PatternFill('solid', fgColor='FFE699')   # Tier S
WIN_FILL   = PatternFill('solid', fgColor='C6E0B4')   # Tier A / WON
LOSE_FILL  = PatternFill('solid', fgColor='F4B6B6')   # LOSE
ACTIVE_FILL= PatternFill('solid', fgColor='BDD7EE')   # active deals
DUPE_FILL  = PatternFill('solid', fgColor='F8CBAD')   # dupes
LINK_FILL  = PatternFill('solid', fgColor='D9D2E9')   # LinkedIn match
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
LINK_FONT = Font(color='0563C1', underline='single')

ACTIVE_STAGES = {'NEW', 'PREPARATION', 'PREPAYMENT_INVOICE', 'EXECUTING', 'FINAL_INVOICE'}


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 36
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def normalize(s):
    if not s:
        return ''
    s = str(s).lower().replace('ё', 'е').replace('"', '').replace("'", '').replace('«', '').replace('»', '')
    s = re.sub(r'\b(ооо|оао|пао|зао|ао|ип|ук|тоо|чп|нко|гк|холдинг|групп|group|llc|ltd|inc|corp|gmbh)\b', '', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


LATIN_TO_CYRILLIC = {
    'fixprice': 'фикс прайс', 'rusagro': 'русагро', 'uralchem': 'уралхим',
    'ingosstrakh': 'ингосстрах', 'invitro': 'инвитро', 'splat': 'сплат',
    'selectel': 'селектел', 'la redoute russia': 'ла редут', 'sanoma learning russia': 'саномо',
}


def fuzzy_match(a, b):
    na, nb = normalize(a), normalize(b)
    if not na or not nb:
        return False
    if na == nb or na in nb or nb in na:
        return True
    ta, tb = set(na.split()), set(nb.split())
    common = {t for t in (ta & tb) if len(t) >= 4}
    if len(common) >= 2:
        return True
    for lat, cyr in LATIN_TO_CYRILLIC.items():
        if (lat in na and cyr in nb) or (lat in nb and cyr in na) \
           or (cyr in na and lat in nb) or (cyr in nb and lat in na):
            return True
    return False


def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace('Z', '+00:00'))
    except Exception:
        return None


def is_active(stage_id):
    if not stage_id:
        return False
    s = stage_id.split(':')[-1] if ':' in stage_id else stage_id
    return s in ACTIVE_STAGES or s.startswith('UC_')


def is_won(stage_id):
    return stage_id and stage_id.split(':')[-1] == 'WON'


def is_lost(stage_id):
    s = stage_id.split(':')[-1] if ':' in (stage_id or '') else (stage_id or '')
    return s in ('LOSE', 'APOLOGY')


def contacts_to_text(contacts):
    return '\n'.join(
        f"{c['name']}"
        f"{' • ' + c['post'] if c.get('post') else ''}"
        f"{' • ' + c['phone'] if c.get('phone') else ''}"
        f"{' • ' + c['email'] if c.get('email') else ''}"
        for c in contacts
    )


def linkedin_summary(matches):
    if not matches:
        return ''
    lines = []
    for m in matches:
        line = m.get('name', '')
        if m.get('role'):
            line += f" • {m['role']}"
        if m.get('linkedin'):
            line += f" • {m['linkedin']}"
        if m.get('reaction'):
            line += f" • [{m['reaction']}]"
        lines.append(line)
    return '\n'.join(lines)


# ========== Build ==========

def build(args):
    extracted = json.load(open(args.input, encoding='utf-8'))
    bd = json.load(open(args.bitrix, encoding='utf-8'))
    by_inn = bd.get('by_inn', {})
    bitrix_base = bd.get('bitrix_base') or args.bitrix_url

    research = {}
    queue_map = {}
    if args.research and os.path.exists(args.research):
        research = json.load(open(args.research, encoding='utf-8'))
    if args.research_queue and os.path.exists(args.research_queue):
        queue = json.load(open(args.research_queue, encoding='utf-8'))
        queue_map = {p['inn']: p for p in queue}

    linkedin = []
    if args.linkedin and os.path.exists(args.linkedin):
        linkedin = json.load(open(args.linkedin, encoding='utf-8'))

    # Group extracted by source for original-sheet augmentation
    by_source = {}
    for r in extracted:
        by_source.setdefault(r['source'], []).append(r)

    # Index linkedin by source-row name (fuzzy)
    linkedin_by_source_name = {}
    for r in extracted:
        if not linkedin:
            break
        matches = []
        for li in linkedin:
            if not li.get('company'):
                continue
            if fuzzy_match(r.get('name', ''), li['company']) or \
               (r.get('ul') and fuzzy_match(r['ul'], li['company'])):
                matches.append(li)
        if matches:
            linkedin_by_source_name[r['name']] = matches

    # Start workbook from source if provided
    if args.source_xlsx and os.path.exists(args.source_xlsx):
        shutil.copyfile(args.source_xlsx, args.output)
        wb = openpyxl.load_workbook(args.output)
        print(f'Loaded source xlsx: {args.source_xlsx}, sheets: {wb.sheetnames}')
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    today = datetime.now()

    # ---------- Augment original sheets ----------
    if args.source_xlsx and os.path.exists(args.source_xlsx):
        for sheet_name, source_records in by_source.items():
            if sheet_name not in wb.sheetnames:
                continue
            augment_original_sheet(wb[sheet_name], source_records, by_inn, linkedin_by_source_name, today)

    # ---------- Sheet: Пересечения ----------
    build_intersections(wb, extracted, by_inn, linkedin_by_source_name, today, research)

    # ---------- Sheet: Готовый обзвон ----------
    build_call_list(wb, extracted, by_inn, linkedin_by_source_name)

    # ---------- Sheet: Активные сделки ----------
    build_active_deals(wb, extracted, by_inn, bitrix_base)

    # ---------- Sheet: Продажи и провалы ----------
    build_won_lost(wb, extracted, by_inn, bitrix_base)

    # ---------- Sheet: Холодняк ----------
    build_cold(wb, extracted, by_inn, linkedin_by_source_name)

    # ---------- Sheet: Дубли источников ----------
    build_dupes(wb, extracted, by_inn)

    # ---------- Sheet: LinkedIn ЛПР ----------
    if linkedin:
        build_linkedin_sheet(wb, extracted, by_inn, linkedin)

    # ---------- Sheet: Реанимация ----------
    build_reanimation(wb, extracted, by_inn, linkedin_by_source_name, today)

    # ---------- Sheet: Сделки — детали ----------
    build_deal_details(wb, by_inn, bitrix_base)

    # ---------- Sheet: Касания — лента ----------
    build_activity_feed(wb, by_inn)

    # ---------- Sheet: Глубокий рисёрч ----------
    if research:
        build_deep_research(wb, research, queue_map)

    # Reorder: put summary sheets at front
    desired_order = [
        'Пересечения', 'Глубокий рисёрч', 'Готовый обзвон',
        'Активные сделки', 'Продажи и провалы', 'Холодняк (нет в базе)',
        'Дубли источников', 'LinkedIn ЛПР', 'Реанимация (90+ дней)',
        'Сделки — детали', 'Касания — лента',
    ]
    existing_order = list(wb.sheetnames)
    new_order = [s for s in desired_order if s in existing_order]
    for s in existing_order:
        if s not in new_order:
            new_order.append(s)
    # Re-order by repeatedly moving
    for i, s in enumerate(new_order):
        idx = wb.sheetnames.index(s)
        if idx != i:
            wb.move_sheet(s, offset=i - idx)

    wb.save(args.output)
    print(f'\n=== Saved: {args.output} ===')


def augment_original_sheet(ws, source_records, by_inn, linkedin_by_source_name, today):
    last_col = ws.max_column
    new_headers = [
        'Match', 'Bitrix компания', 'Bitrix URL', 'Все продукты (категории)', 'Стадии сделок',
        'Сделок всего', 'Контактов в Bitrix', 'Касаний всего', 'Последнее касание',
        'Дней с касания', 'Менеджер(ы)', 'Контакты Bitrix', 'LinkedIn ЛПР',
    ]
    for i, h in enumerate(new_headers, start=1):
        ws.cell(row=1, column=last_col + i, value=h)
    style_header(ws, 1, last_col + len(new_headers))
    widths = [14, 36, 36, 30, 30, 12, 14, 14, 16, 14, 28, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(last_col + i)].width = w

    for i, rec in enumerate(source_records, start=rec_first_row(ws)):
        inn = rec.get('inn', '')
        info = by_inn.get(inn) if inn else None
        col = last_col
        li_matches = linkedin_by_source_name.get(rec.get('name', ''), [])

        if info:
            col += 1; ws.cell(row=i, column=col, value='ДА')
            col += 1; ws.cell(row=i, column=col, value=' / '.join(info['titles']))
            col += 1
            if info.get('urls'):
                cell = ws.cell(row=i, column=col, value='Открыть в Bitrix')
                cell.hyperlink = info['urls'][0]
                cell.font = LINK_FONT
            col += 1; ws.cell(row=i, column=col, value=', '.join(info.get('products') or []))
            col += 1; ws.cell(row=i, column=col, value=', '.join(info.get('stages_used') or []))
            col += 1; ws.cell(row=i, column=col, value=len(info.get('deals') or []))
            col += 1; ws.cell(row=i, column=col, value=len(info.get('contacts') or []))
            col += 1; ws.cell(row=i, column=col, value=info.get('touch_count', 0))
            col += 1; ws.cell(row=i, column=col, value=info.get('last_touch') or '')
            col += 1
            if info.get('last_touch'):
                lt = parse_dt(info['last_touch'])
                if lt:
                    days = (datetime.now() - lt.replace(tzinfo=None)).days
                    ws.cell(row=i, column=col, value=days)
            col += 1; ws.cell(row=i, column=col, value=', '.join(info.get('managers') or []))
            col += 1
            ct = contacts_to_text(info.get('contacts') or [])
            ws.cell(row=i, column=col, value=ct).alignment = WRAP
            col += 1
            ws.cell(row=i, column=col, value=linkedin_summary(li_matches)).alignment = WRAP

            row_fill = HOT_FILL if info.get('touch_count', 0) > 0 else MATCH_FILL
            for c in range(1, ws.max_column + 1):
                cur = ws.cell(row=i, column=c)
                if cur.fill.fgColor.value in (None, '00000000'):
                    cur.fill = row_fill
        else:
            col += 1; ws.cell(row=i, column=col, value='-')
            if li_matches:
                col_li = last_col + len(new_headers)
                ws.cell(row=i, column=col_li, value=linkedin_summary(li_matches)).alignment = WRAP
                ws.cell(row=i, column=col_li).fill = LINK_FILL


def rec_first_row(ws):
    """Return first data row index in original sheet (assume row 1 = headers)."""
    return 2


# ========== Sheet builders ==========

def build_intersections(wb, extracted, by_inn, linkedin_by_source_name, today, research):
    if 'Пересечения' in wb.sheetnames:
        del wb['Пересечения']
    ws = wb.create_sheet('Пересечения', 0)
    headers = [
        'Источник', 'ОРГ (из файла)', 'ИНН', 'ЮЛ (из файла)', 'Bitrix компания', 'Bitrix ID',
        'Карточка в Bitrix', 'Сделок', 'Активных', 'Контактов', 'Касаний', 'Последнее', 'Дней',
        'Продукты', 'Стадии', 'Менеджеры', 'Контакты Bitrix', 'LinkedIn ЛПР', 'Сводка касаний',
    ]
    if research:
        headers.append('Свежий рисёрч (Perplexity)')
    ws.append(headers)
    widths = [10, 32, 13, 32, 32, 10, 18, 8, 9, 9, 9, 12, 8, 28, 28, 22, 50, 50, 60]
    if research:
        widths.append(80)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    seen = set()
    rows = []
    for r in extracted:
        inn = r.get('inn')
        if not inn or inn in seen:
            continue
        info = by_inn.get(inn)
        if not info:
            continue
        seen.add(inn)
        rows.append((r['source'], r, info))

    rows.sort(key=lambda x: (
        -1 if x[2]['touch_count'] > 0 else 0,
        -1 if x[2]['has_deals'] else 0,
        -1 if x[2]['has_contacts'] else 0,
        x[1].get('name') or '',
    ))

    for src_label, rec, info in rows:
        contacts_text = contacts_to_text(info.get('contacts') or [])
        touches = []
        for d in (info.get('deals') or [])[:10]:
            amt = d.get('OPPORTUNITY') or 0
            try:
                amt_str = f"{float(amt):,.0f} {d.get('CURRENCY_ID', 'RUB')}".replace(',', ' ')
            except Exception:
                amt_str = str(amt)
            date = (d.get('DATE_CREATE') or '')[:10]
            touches.append(f"[deal] {date} • {d.get('_category_name', '')} • {d.get('_stage_name', '')} • {amt_str}")
        for a in (info.get('activities') or [])[:10]:
            check = 'V' if a.get('completed') else 'o'
            touches.append(f"{check} {a.get('date', '')} • {a.get('type', '')} • {(a.get('subject') or '')[:60]}")
        touch_summary = '\n'.join(touches)

        days = ''
        if info.get('last_touch'):
            lt = parse_dt(info['last_touch'])
            if lt:
                days = (today - lt.replace(tzinfo=None)).days

        active_count = sum(1 for d in (info.get('deals') or []) if is_active(d.get('STAGE_ID', '')))
        li_matches = linkedin_by_source_name.get(rec.get('name', ''), [])
        li_text = linkedin_summary(li_matches)

        row_data = [
            src_label, rec.get('name', ''), rec.get('inn', ''), rec.get('ul', ''),
            ' / '.join(info.get('titles') or []), ', '.join(info.get('bitrix_ids') or []),
            '',
            len(info.get('deals') or []), active_count, len(info.get('contacts') or []),
            info.get('touch_count', 0), info.get('last_touch') or '', days,
            ', '.join(info.get('products') or []), ', '.join(info.get('stages_used') or []),
            ', '.join(info.get('managers') or []),
            contacts_text, li_text, touch_summary,
        ]
        if research:
            r_info = research.get(rec.get('inn', ''), {})
            if r_info.get('status') in ('ok', 'cached'):
                row_data.append(r_info.get('text', ''))
            else:
                row_data.append('')

        ws.append(row_data)
        r_idx = ws.max_row

        if info.get('touch_count', 0) > 0:
            fill = HOT_FILL
        elif info.get('has_deals') or info.get('has_contacts'):
            fill = MATCH_FILL
        else:
            fill = None
        if fill:
            for c in range(1, len(row_data) + 1):
                ws.cell(row=r_idx, column=c).fill = fill

        if info.get('urls'):
            cell = ws.cell(row=r_idx, column=7, value='Открыть в Bitrix')
            cell.hyperlink = info['urls'][0]
            cell.font = LINK_FONT

        wrap_cols = [17, 18, 19]
        if research:
            wrap_cols.append(20)
        for col in wrap_cols:
            ws.cell(row=r_idx, column=col).alignment = WRAP

        line_count = max(
            len(contacts_text.split('\n')) if contacts_text else 1,
            len(touch_summary.split('\n')) if touch_summary else 1,
            len(li_text.split('\n')) if li_text else 1,
            (row_data[-1].count('\n') + 1) if research and row_data[-1] else 1,
        )
        ws.row_dimensions[r_idx].height = max(20, min(300, line_count * 14))

    # Title block
    n_headers = len(headers)
    ws.insert_rows(1, amount=4)
    ws.cell(row=1, column=1, value='Пересечения списка × CRM')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_headers)
    ws.cell(row=2, column=1, value=(
        f'Всего матчей по ИНН: {len(rows)} | '
        f'Со сделками: {sum(1 for _,_,i in rows if i["has_deals"])} | '
        f'С контактами: {sum(1 for _,_,i in rows if i["has_contacts"])} | '
        f'С касаниями: {sum(1 for _,_,i in rows if i["touch_count"]>0)}'
    ))
    ws.cell(row=2, column=1).font = Font(italic=True, color='595959')
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=n_headers)
    ws.cell(row=3, column=1, value='жёлтый — активные касания, зелёный — в базе без касаний; строки отсортированы по приоритету')
    ws.cell(row=3, column=1).font = Font(italic=True, color='595959', size=9)
    ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=n_headers)
    for r in range(1, 8):
        if ws.cell(row=r, column=1).value == 'Источник':
            style_header(ws, r, n_headers)
            ws.freeze_panes = ws.cell(row=r + 1, column=1)
            ws.auto_filter.ref = f'A{r}:{get_column_letter(n_headers)}{ws.max_row}'
            break


def build_call_list(wb, extracted, by_inn, linkedin_by_source_name):
    if 'Готовый обзвон' in wb.sheetnames:
        del wb['Готовый обзвон']
    ws = wb.create_sheet('Готовый обзвон')
    headers = [
        'Приоритет', 'Источник', 'Компания', 'ИНН', 'Bitrix компания', 'Карточка',
        'ФИО контакта', 'Должность', 'Телефон', 'Email', 'LinkedIn',
        'Источник контакта', 'Менеджер CRM', 'Продукты в работе', 'Последнее касание', 'Комментарий',
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [10, 10, 30, 13, 30, 14, 28, 22, 22, 28, 40, 14, 22, 24, 14, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    call_rows = []
    by_inn_rec = {}
    for r in extracted:
        if r.get('inn'):
            by_inn_rec.setdefault(r['inn'], (r['source'], r))

    # Source 1: Bitrix contacts with phone/email
    for inn, info in by_inn.items():
        src_label, rec = by_inn_rec.get(inn, ('', {}))
        for c in info.get('contacts', []):
            if not (c.get('phone') or c.get('email')):
                continue
            priority = 'A' if info.get('touch_count', 0) > 0 else 'B'
            call_rows.append({
                'priority': priority, 'source': src_label,
                'company': rec.get('name', ''), 'inn': inn,
                'bitrix_title': ' / '.join(info.get('titles') or []),
                'bitrix_url': (info.get('urls') or [''])[0],
                'name': c.get('name', ''), 'post': c.get('post', ''),
                'phone': c.get('phone', ''), 'email': c.get('email', ''),
                'linkedin': '',
                'origin': 'Bitrix',
                'manager': ', '.join(info.get('managers') or []),
                'products': ', '.join(info.get('products') or []),
                'last_touch': info.get('last_touch') or '',
                'comment': rec.get('last_call_comment', ''),
            })

    # Source 2: LinkedIn contacts
    for r in extracted:
        li_matches = linkedin_by_source_name.get(r.get('name', ''), [])
        if not li_matches:
            continue
        info = by_inn.get(r.get('inn', ''))
        for li in li_matches:
            priority = 'C'
            if info and info.get('touch_count', 0) > 0:
                priority = 'A'
            elif info:
                priority = 'B'
            call_rows.append({
                'priority': priority, 'source': r.get('source', ''),
                'company': r.get('name', ''), 'inn': r.get('inn', ''),
                'bitrix_title': ' / '.join(info.get('titles') or []) if info else '',
                'bitrix_url': (info.get('urls') or [''])[0] if info else '',
                'name': li.get('name', ''), 'post': li.get('role', ''),
                'phone': '', 'email': r.get('email', '') or '',
                'linkedin': li.get('linkedin', ''),
                'origin': 'LinkedIn',
                'manager': ', '.join(info.get('managers') or []) if info else '',
                'products': ', '.join(info.get('products') or []) if info else '',
                'last_touch': info.get('last_touch') if info else '',
                'comment': li.get('reaction') or li.get('notes') or '',
            })

    # Source 3: Bitrix companies — surface phone/email from source row if no named contact
    for inn, info in by_inn.items():
        if any(c.get('phone') or c.get('email') for c in info.get('contacts', [])):
            continue
        src_label, rec = by_inn_rec.get(inn, ('', {}))
        src_phone = rec.get('phone', '')
        src_email = rec.get('email', '')
        if not (src_phone or src_email):
            continue
        priority = 'A' if info.get('touch_count', 0) > 0 else 'B'
        call_rows.append({
            'priority': priority, 'source': src_label,
            'company': rec.get('name', ''), 'inn': inn,
            'bitrix_title': ' / '.join(info.get('titles') or []),
            'bitrix_url': (info.get('urls') or [''])[0],
            'name': '(общий — звонить в компанию)', 'post': '',
            'phone': src_phone, 'email': src_email, 'linkedin': '',
            'origin': 'Источник',
            'manager': ', '.join(info.get('managers') or []),
            'products': ', '.join(info.get('products') or []),
            'last_touch': info.get('last_touch') or '',
            'comment': rec.get('last_call_comment', ''),
        })

    call_rows.sort(key=lambda x: (x['priority'], x['company']))
    for cr in call_rows:
        ws.append([
            cr['priority'], cr['source'], cr['company'], cr['inn'], cr['bitrix_title'], '',
            cr['name'], cr['post'], cr['phone'], cr['email'], cr['linkedin'],
            cr['origin'], cr['manager'], cr['products'], cr['last_touch'], cr['comment'],
        ])
        rr = ws.max_row
        if cr['bitrix_url']:
            cell = ws.cell(row=rr, column=6, value='Bitrix')
            cell.hyperlink = cr['bitrix_url']
            cell.font = LINK_FONT
        if cr['linkedin']:
            cell = ws.cell(row=rr, column=11)
            cell.hyperlink = cr['linkedin']
            cell.font = LINK_FONT
        fill = HOT_FILL if cr['priority'] == 'A' else (MATCH_FILL if cr['priority'] == 'B' else LINK_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = fill
    ws.auto_filter.ref = ws.dimensions


def build_active_deals(wb, extracted, by_inn, bitrix_base):
    if 'Активные сделки' in wb.sheetnames:
        del wb['Активные сделки']
    ws = wb.create_sheet('Активные сделки')
    headers = ['ИНН', 'Компания (из файла)', 'Компания (Bitrix)', 'ID сделки', 'Сделка', 'Карточка',
               'Продукт', 'Стадия', 'Сумма', 'Создана', 'Менеджер', 'Контакты Bitrix']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [13, 30, 30, 10, 38, 16, 25, 24, 14, 14, 24, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    inn_to_src = {}
    for r in extracted:
        if r.get('inn') and r['inn'] not in inn_to_src:
            inn_to_src[r['inn']] = r.get('name', '')

    for inn, info in by_inn.items():
        for d in info.get('deals', []):
            if not is_active(d.get('STAGE_ID', '')):
                continue
            cid = d.get('_company_id', '')
            titles = info.get('titles') or []
            bitrix_ids = info.get('bitrix_ids') or []
            comp_title = next((t for t, bid in zip(titles, bitrix_ids) if bid == cid),
                              titles[0] if titles else '')
            try:
                opp = float(d.get('OPPORTUNITY') or 0)
            except Exception:
                opp = 0
            contacts_text = '\n'.join(
                f"{c['name']}{' • ' + c['post'] if c.get('post') else ''}"
                f"{' • ' + c['phone'] if c.get('phone') else ''}"
                f"{' • ' + c['email'] if c.get('email') else ''}"
                for c in info.get('contacts', []) if c.get('company_id') == cid
            )
            ws.append([
                inn, inn_to_src.get(inn, ''), comp_title, d['ID'], d.get('TITLE', ''), '',
                d.get('_category_name', ''), d.get('_stage_name', ''), opp,
                (d.get('DATE_CREATE') or '')[:10], d.get('_manager', ''), contacts_text,
            ])
            rr = ws.max_row
            cell = ws.cell(row=rr, column=6, value='Открыть')
            cell.hyperlink = f'{bitrix_base}/crm/deal/details/{d["ID"]}/'
            cell.font = LINK_FONT
            for c in range(1, len(headers) + 1):
                ws.cell(row=rr, column=c).fill = ACTIVE_FILL
            ws.cell(row=rr, column=12).alignment = WRAP
    ws.auto_filter.ref = ws.dimensions


def build_won_lost(wb, extracted, by_inn, bitrix_base):
    if 'Продажи и провалы' in wb.sheetnames:
        del wb['Продажи и провалы']
    ws = wb.create_sheet('Продажи и провалы')
    headers = ['Результат', 'ИНН', 'Компания (из файла)', 'Компания (Bitrix)', 'Сделка', 'Карточка',
               'Продукт', 'Стадия', 'Сумма', 'Создана', 'Закрыта', 'Менеджер']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [10, 13, 30, 30, 38, 16, 25, 25, 14, 14, 14, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    inn_to_src = {r['inn']: r.get('name', '') for r in extracted if r.get('inn')}
    for inn, info in by_inn.items():
        for d in info.get('deals', []):
            sid = d.get('STAGE_ID', '')
            if is_won(sid):
                result, fill = 'УСПЕХ', WIN_FILL
            elif is_lost(sid):
                result, fill = 'ПРОВАЛ', LOSE_FILL
            else:
                continue
            cid = d.get('_company_id', '')
            titles = info.get('titles') or []
            bitrix_ids = info.get('bitrix_ids') or []
            comp_title = next((t for t, bid in zip(titles, bitrix_ids) if bid == cid),
                              titles[0] if titles else '')
            try:
                opp = float(d.get('OPPORTUNITY') or 0)
            except Exception:
                opp = 0
            ws.append([
                result, inn, inn_to_src.get(inn, ''), comp_title, d.get('TITLE', ''), '',
                d.get('_category_name', ''), d.get('_stage_name', ''), opp,
                (d.get('DATE_CREATE') or '')[:10], (d.get('CLOSEDATE') or '')[:10],
                d.get('_manager', ''),
            ])
            rr = ws.max_row
            cell = ws.cell(row=rr, column=6, value='Открыть')
            cell.hyperlink = f'{bitrix_base}/crm/deal/details/{d["ID"]}/'
            cell.font = LINK_FONT
            for c in range(1, len(headers) + 1):
                ws.cell(row=rr, column=c).fill = fill
    ws.auto_filter.ref = ws.dimensions


def build_cold(wb, extracted, by_inn, linkedin_by_source_name):
    if 'Холодняк (нет в базе)' in wb.sheetnames:
        del wb['Холодняк (нет в базе)']
    ws = wb.create_sheet('Холодняк (нет в базе)')
    headers = ['Источник', 'ОРГ', 'ИНН', 'ЮЛ', 'Выручка 2024 (RUB)', 'Выручка 2023 (RUB)',
               'Телефон', 'Email', 'Сайт', 'Отрасль', 'LinkedIn ЛПР', 'Сегмент', 'Комментарий']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [10, 32, 13, 32, 18, 18, 22, 28, 28, 24, 50, 22, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    cold = []
    for r in extracted:
        if r.get('inn') in by_inn:
            continue
        rev = r.get('rev2024') or 0
        try:
            rev_f = float(rev) if rev else 0
        except Exception:
            rev_f = 0
        rev2023 = r.get('rev2023') or 0
        try:
            rev2023_f = float(rev2023) if rev2023 else 0
        except Exception:
            rev2023_f = 0
        li_matches = linkedin_by_source_name.get(r.get('name', ''), [])
        cold.append({
            'source': r.get('source', ''), 'name': r.get('name', ''),
            'inn': r.get('inn', ''), 'ul': r.get('ul', ''),
            'rev2024': rev_f, 'rev2023': rev2023_f,
            'phone': r.get('phone', ''), 'email': r.get('email', ''),
            'site': r.get('site', ''),
            'industry': r.get('industry', '') or r.get('segment', ''),
            'li': linkedin_summary(li_matches),
            'segment': r.get('segment', ''),
            'comment': r.get('last_call_comment', ''),
        })
    cold.sort(key=lambda x: -x['rev2024'])
    for c in cold:
        ws.append([
            c['source'], c['name'], c['inn'], c['ul'], c['rev2024'], c['rev2023'],
            c['phone'], c['email'], c['site'], c['industry'], c['li'], c['segment'], c['comment'],
        ])
        rr = ws.max_row
        if c['li']:
            for col in range(1, len(headers) + 1):
                ws.cell(row=rr, column=col).fill = LINK_FILL
            ws.cell(row=rr, column=11).alignment = WRAP
        ws.cell(row=rr, column=5).number_format = '#,##0'
        ws.cell(row=rr, column=6).number_format = '#,##0'
    ws.auto_filter.ref = ws.dimensions


def build_dupes(wb, extracted, by_inn):
    # Find INNs appearing in >1 source
    inn_to_sources = {}
    for r in extracted:
        inn = r.get('inn')
        if not inn:
            continue
        inn_to_sources.setdefault(inn, []).append(r)
    dupes = {inn: lst for inn, lst in inn_to_sources.items() if len({r['source'] for r in lst}) >= 2}
    if not dupes:
        return

    if 'Дубли источников' in wb.sheetnames:
        del wb['Дубли источников']
    ws = wb.create_sheet('Дубли источников')
    headers = ['ИНН', 'ОРГ (1)', 'ОРГ (2)', 'ЮЛ', 'В Bitrix?', 'Касаний', 'Bitrix URL']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [13, 32, 32, 32, 12, 10, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for inn in sorted(dupes):
        rs = dupes[inn]
        r1 = rs[0]
        r2 = rs[1] if len(rs) > 1 else {}
        info = by_inn.get(inn)
        in_bitrix = 'ДА' if info else 'нет'
        touches = info.get('touch_count', '') if info else ''
        url = (info.get('urls') or [''])[0] if info else ''
        ws.append([inn, r1.get('name', ''), r2.get('name', ''),
                   r1.get('ul', ''), in_bitrix, touches, ''])
        rr = ws.max_row
        if url:
            cell = ws.cell(row=rr, column=7, value='Открыть')
            cell.hyperlink = url
            cell.font = LINK_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = DUPE_FILL
    ws.auto_filter.ref = ws.dimensions


def build_linkedin_sheet(wb, extracted, by_inn, linkedin):
    if 'LinkedIn ЛПР' in wb.sheetnames:
        del wb['LinkedIn ЛПР']
    ws = wb.create_sheet('LinkedIn ЛПР')
    headers = ['Компания (LinkedIn)', 'Совпадение из файла', 'Источник', 'ИНН',
               'В Bitrix?', 'ФИО', 'Должность', 'LinkedIn', 'Индустрия',
               'Инвайт', 'Принял', 'Ответ']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [28, 30, 12, 13, 11, 26, 24, 40, 22, 14, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    src_index = {}
    for r in extracted:
        src_index.setdefault(r.get('name', ''), (r.get('source', ''), r))

    for li in linkedin:
        if not li.get('company'):
            continue
        matched = None
        for src_name, (src_label, src_rec) in src_index.items():
            if fuzzy_match(li['company'], src_name) or fuzzy_match(li['company'], src_rec.get('ul', '')):
                matched = (src_label, src_rec)
                break
        src_label = matched[0] if matched else ''
        src_rec = matched[1] if matched else None
        info = by_inn.get(src_rec['inn']) if src_rec and src_rec.get('inn') else None
        in_bitrix = 'ДА' if info else 'нет'

        ws.append([
            li['company'],
            src_rec.get('name', '') if src_rec else '— нет совпадения —',
            src_label, src_rec.get('inn', '') if src_rec else '', in_bitrix,
            li.get('name', ''), li.get('role', ''), '', li.get('segment', ''),
            li.get('reaction', ''), li.get('screen', ''), li.get('notes', ''),
        ])
        rr = ws.max_row
        if li.get('linkedin'):
            cell = ws.cell(row=rr, column=8, value=li['linkedin'])
            cell.hyperlink = li['linkedin']
            cell.font = LINK_FONT
        if matched:
            for c in range(1, len(headers) + 1):
                ws.cell(row=rr, column=c).fill = LINK_FILL
    ws.auto_filter.ref = ws.dimensions


def build_reanimation(wb, extracted, by_inn, linkedin_by_source_name, today):
    if 'Реанимация (90+ дней)' in wb.sheetnames:
        del wb['Реанимация (90+ дней)']
    ws = wb.create_sheet('Реанимация (90+ дней)')
    headers = ['ИНН', 'Компания', 'Bitrix компания', 'Карточка', 'Дней с касания',
               'Последнее касание', 'Продукты', 'Стадии', 'Менеджер', 'Контакты Bitrix', 'LinkedIn ЛПР']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [13, 30, 30, 14, 14, 14, 25, 25, 22, 50, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    inn_to_src = {r['inn']: r.get('name', '') for r in extracted if r.get('inn')}
    rean_rows = []
    for inn, info in by_inn.items():
        if info.get('touch_count', 0) == 0:
            continue
        if not info.get('last_touch'):
            continue
        lt = parse_dt(info['last_touch'])
        if not lt:
            continue
        days = (today - lt.replace(tzinfo=None)).days
        if days < 90:
            continue
        src_name = inn_to_src.get(inn, '')
        li = linkedin_by_source_name.get(src_name, [])
        rean_rows.append((days, inn, src_name, info, li))

    rean_rows.sort(key=lambda x: -x[0])
    for days, inn, src_name, info, li in rean_rows:
        contacts_text = contacts_to_text(info.get('contacts') or [])
        ws.append([
            inn, src_name, ' / '.join(info.get('titles') or []), '',
            days, info.get('last_touch', ''),
            ', '.join(info.get('products') or []), ', '.join(info.get('stages_used') or []),
            ', '.join(info.get('managers') or []), contacts_text, linkedin_summary(li),
        ])
        rr = ws.max_row
        if info.get('urls'):
            cell = ws.cell(row=rr, column=4, value='Открыть')
            cell.hyperlink = info['urls'][0]
            cell.font = LINK_FONT
        ws.cell(row=rr, column=10).alignment = WRAP
        ws.cell(row=rr, column=11).alignment = WRAP
    ws.auto_filter.ref = ws.dimensions


def build_deal_details(wb, by_inn, bitrix_base):
    if 'Сделки — детали' in wb.sheetnames:
        del wb['Сделки — детали']
    ws = wb.create_sheet('Сделки — детали')
    headers = ['ИНН', 'Компания (Bitrix)', 'Bitrix ID', 'ID сделки', 'Сделка', 'Карточка',
               'Категория', 'Стадия', 'Сумма', 'Валюта', 'Создана', 'Закрыта', 'Закрыта?', 'Менеджер']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [13, 30, 10, 10, 38, 18, 25, 22, 14, 8, 14, 14, 10, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for inn, info in by_inn.items():
        for d in info.get('deals', []):
            cid = d.get('_company_id', '')
            titles = info.get('titles') or []
            bitrix_ids = info.get('bitrix_ids') or []
            comp_title = next((t for t, bid in zip(titles, bitrix_ids) if bid == cid),
                              titles[0] if titles else '')
            try:
                opp = float(d.get('OPPORTUNITY') or 0)
            except Exception:
                opp = 0
            ws.append([
                inn, comp_title, cid, d['ID'], d.get('TITLE', ''), '',
                d.get('_category_name', ''), d.get('_stage_name', ''), opp, d.get('CURRENCY_ID', ''),
                (d.get('DATE_CREATE') or '')[:10], (d.get('CLOSEDATE') or '')[:10],
                'Да' if d.get('CLOSED') == 'Y' else 'Нет', d.get('_manager', ''),
            ])
            rr = ws.max_row
            cell = ws.cell(row=rr, column=6, value='Открыть')
            cell.hyperlink = f'{bitrix_base}/crm/deal/details/{d["ID"]}/'
            cell.font = LINK_FONT
    ws.auto_filter.ref = ws.dimensions


def build_activity_feed(wb, by_inn):
    if 'Касания — лента' in wb.sheetnames:
        del wb['Касания — лента']
    ws = wb.create_sheet('Касания — лента')
    headers = ['Дата', 'ИНН', 'Компания', 'Тип касания', 'Тема', 'Источник', 'Завершено?']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [12, 13, 32, 12, 50, 18, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    flat = []
    for inn, info in by_inn.items():
        if not info.get('activities'):
            continue
        title = (info.get('titles') or [''])[0]
        for a in info['activities']:
            flat.append((a.get('date') or '', inn, title, a.get('type', ''),
                         a.get('subject') or '', a.get('source', ''),
                         'Да' if a.get('completed') else 'Нет'))
    flat.sort(key=lambda x: x[0], reverse=True)
    for row in flat:
        ws.append(row)
    ws.auto_filter.ref = ws.dimensions


def build_deep_research(wb, research, queue_map):
    if 'Глубокий рисёрч' in wb.sheetnames:
        del wb['Глубокий рисёрч']
    ws = wb.create_sheet('Глубокий рисёрч')
    headers = ['Tier', 'Компания', 'ИНН', 'ЮЛ', 'Отрасль', 'Причина в очереди', 'Досье (Perplexity)']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    widths = [6, 32, 13, 32, 22, 45, 100]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    sorted_inns = sorted(research.keys(), key=lambda inn: (
        queue_map.get(inn, {}).get('tier', 'Z'),
        queue_map.get(inn, {}).get('name', ''),
    ))
    for inn in sorted_inns:
        r = research[inn]
        if r.get('status') not in ('ok', 'cached'):
            continue
        p = queue_map.get(inn, {})
        text = r.get('text', '')
        ws.append([
            p.get('tier', ''), p.get('name', r.get('name', '')), inn,
            p.get('ul', ''), p.get('industry', ''), p.get('reason', ''), text,
        ])
        rr = ws.max_row
        tier = p.get('tier', '')
        fill = HOT_FILL if tier == 'S' else (MATCH_FILL if tier == 'B' else WIN_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = fill
        cell = ws.cell(row=rr, column=7)
        cell.alignment = WRAP
        n_lines = max(8, min(40, text.count('\n') + 1))
        ws.row_dimensions[rr].height = max(80, n_lines * 14)
    ws.auto_filter.ref = ws.dimensions


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True, help='extracted.json')
    ap.add_argument('--bitrix', required=True, help='bitrix_data.json')
    ap.add_argument('--output', required=True, help='enriched.xlsx output')
    ap.add_argument('--source-xlsx', default='', help='original xlsx — keep its sheets and augment them')
    ap.add_argument('--research', default='', help='research/results.json (optional)')
    ap.add_argument('--research-queue', default='', help='research/queue.json (optional)')
    ap.add_argument('--linkedin', default='', help='linkedin.json (optional)')
    ap.add_argument('--bitrix-url', default='https://we.company.example')
    args = ap.parse_args()
    build(args)


if __name__ == '__main__':
    main()
