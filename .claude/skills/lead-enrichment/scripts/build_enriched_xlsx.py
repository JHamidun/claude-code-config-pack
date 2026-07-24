# -*- coding: utf-8 -*-
"""Build the multi-sheet enriched workbook. Generalized port of the proven build_xlsx.

Usage:  python build_enriched_xlsx.py --list list.json --agg aggregated.json \
            [--firmographics firmographics.json] [--research research/results.json] \
            [--linkedin linkedin.json] [--out Enriched.xlsx] [--base https://we.company.example] \
            [--days-reanimate 90]

Sheets (see references/excel-schema.md): Пересечения (summary) · Готовый обзвон ·
Активные сделки · Продажи и провалы · Холодняк (нет в базе) · Реанимация (90+ дней) ·
LinkedIn ЛПР · Глубокий рисёрч · Касания — лента. Each: freeze_panes + auto_filter + color.
"""
import sys, io, json, argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TIER_FILL = {'S': PatternFill('solid', fgColor='C6EFCE'), 'A': PatternFill('solid', fgColor='FFEB9C'),
             'B': PatternFill('solid', fgColor='F2F2F2')}
WRAP = Alignment(wrap_text=True, vertical='top')

def style_header(ws):
    for c in ws[1]:
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def widths(ws, spec):
    for i, w in enumerate(spec, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def stage_kind(name):
    n = (name or '').lower()
    if any(k in n for k in ('провал', 'fail', 'отказ', 'не реализ')):
        return 'lost'
    if any(k in n for k in ('успешн', 'won', 'оплач', 'выигр', 'сделка заключ')):
        return 'won'
    return 'active'

def days_since(date_str):
    d = None
    try:
        d = datetime.fromisoformat(str(date_str))
    except Exception:
        return None
    return (datetime.now() - d).days

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--list', required=True); ap.add_argument('--agg', required=True)
    ap.add_argument('--firmographics'); ap.add_argument('--research'); ap.add_argument('--linkedin')
    ap.add_argument('--out', default='Enriched.xlsx'); ap.add_argument('--base', default='https://we.company.example')
    ap.add_argument('--days-reanimate', type=int, default=90)
    a = ap.parse_args()

    rows = json.load(open(a.list, encoding='utf-8'))
    agg = {k: v for k, v in json.load(open(a.agg, encoding='utf-8')).items() if v}
    firm = json.load(open(a.firmographics, encoding='utf-8')) if a.firmographics else {}
    research = json.load(open(a.research, encoding='utf-8')) if a.research else {}
    linkedin = json.load(open(a.linkedin, encoding='utf-8')) if a.linkedin else []
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    # 1. Пересечения (summary)
    ws = wb.create_sheet('Пересечения')
    matched = len(agg); with_deals = sum(1 for v in agg.values() if v['has_deals'])
    cold = [r for r in rows if str(r.get('inn') or '') not in agg]
    ws.append(['Сводка обогащения']); ws.append([f'Всего в списке: {len(rows)}'])
    ws.append([f'Матчей с Bitrix по ИНН/имени: {matched}']); ws.append([f'Со сделками: {with_deals}'])
    ws.append([f'Холодняк (нет в базе): {len(cold)}'])
    ws.append([f'Дата: {datetime.now():%Y-%m-%d}'])
    ws['A1'].font = Font(bold=True, size=14); widths(ws, [60])

    # 2. Готовый обзвон (Tier S/A priority with contacts)
    ws = wb.create_sheet('Готовый обзвон')
    h = ['Приоритет', 'Компания', 'ИНН', 'ФИО контакта', 'Должность', 'Телефон', 'Email',
         'LinkedIn', 'Менеджер Company', 'Продукты в работе', 'Последнее касание', 'Источник']
    ws.append(h)
    for inn, info in sorted(agg.items(), key=lambda kv: 0 if kv[1]['touch_count'] > 0 else 1):
        tier = 'S' if info['touch_count'] > 0 else 'B'
        comp = info['titles'][0] if info['titles'] else inn
        contacts = info['contacts'] or [{}]
        for c in contacts:
            r = [tier, comp, inn, c.get('name', ''), c.get('post', ''), c.get('phone', ''),
                 c.get('email', ''), '', '; '.join(info['managers']), '; '.join(info['products']),
                 info['last_touch'], 'Bitrix']
            ws.append(r); ws.cell(ws.max_row, 1).fill = TIER_FILL.get(tier, TIER_FILL['B'])
    for r in cold[:1000]:
        ws.append(['A', r.get('name', ''), r.get('inn', ''), '', '', r.get('phone', ''),
                   r.get('email', ''), '', '', '', '', 'Список (холодняк)'])
        ws.cell(ws.max_row, 1).fill = TIER_FILL['A']
    widths(ws, [10, 34, 14, 22, 22, 22, 26, 28, 18, 26, 14, 16]); style_header(ws)

    # 3. Активные сделки
    ws = wb.create_sheet('Активные сделки')
    ws.append(['ИНН', 'Компания', 'ID сделки', 'Сделка', 'Продукт', 'Стадия', 'Сумма', 'Создана', 'Менеджер'])
    for inn, info in agg.items():
        for d in info['deals']:
            if stage_kind(d.get('_stage_name')) == 'active':
                ws.append([inn, info['titles'][0] if info['titles'] else '', d.get('ID'), d.get('TITLE', ''),
                           d.get('_category_name'), d.get('_stage_name'), d.get('OPPORTUNITY', ''),
                           (d.get('DATE_CREATE') or '')[:10], d.get('_manager', '')])
    widths(ws, [14, 30, 12, 28, 22, 24, 12, 12, 18]); style_header(ws)

    # 4. Продажи и провалы
    ws = wb.create_sheet('Продажи и провалы')
    ws.append(['Результат', 'ИНН', 'Компания', 'Сделка', 'Продукт', 'Стадия', 'Сумма', 'Менеджер'])
    for inn, info in agg.items():
        for d in info['deals']:
            k = stage_kind(d.get('_stage_name'))
            if k in ('won', 'lost'):
                ws.append(['ПРОДАЖА' if k == 'won' else 'ПРОВАЛ', inn,
                           info['titles'][0] if info['titles'] else '', d.get('TITLE', ''),
                           d.get('_category_name'), d.get('_stage_name'), d.get('OPPORTUNITY', ''), d.get('_manager', '')])
    widths(ws, [12, 14, 30, 28, 22, 24, 12, 18]); style_header(ws)

    # 5. Холодняк (нет в базе)
    ws = wb.create_sheet('Холодняк (нет в базе)')
    ws.append(['Компания', 'ИНН', 'ЮЛ', 'Выручка 2024', 'Телефон', 'Email', 'Сайт', 'Гендир', 'Сегмент'])
    for r in sorted(cold, key=lambda x: -(float(x.get('rev2024') or 0) if str(x.get('rev2024') or '').replace('.', '').isdigit() else 0)):
        f = firm.get(str(r.get('inn') or ''), {})
        ws.append([r.get('name', ''), r.get('inn', ''), r.get('ul', '') or f.get('name_full', ''),
                   r.get('rev2024') or f.get('income', ''), r.get('phone', ''), r.get('email', ''),
                   r.get('site', ''), f.get('director', ''), r.get('segment', '')])
    widths(ws, [34, 14, 34, 16, 20, 24, 20, 26, 16]); style_header(ws)

    # 6. Реанимация (N+ дней)
    ws = wb.create_sheet(f'Реанимация ({a.days_reanimate}+ дней)')
    ws.append(['ИНН', 'Компания', 'Дней с касания', 'Последнее касание', 'Продукты', 'Менеджер', 'Контакты'])
    rean = []
    for inn, info in agg.items():
        ds = days_since(info['last_touch'])
        if ds is not None and ds >= a.days_reanimate:
            rean.append((ds, inn, info))
    for ds, inn, info in sorted(rean, reverse=True):
        cstr = '; '.join(f"{c.get('name','')} • {c.get('phone','')} • {c.get('email','')}" for c in info['contacts'][:3])
        ws.append([inn, info['titles'][0] if info['titles'] else '', ds, info['last_touch'],
                   '; '.join(info['products']), '; '.join(info['managers']), cstr])
    widths(ws, [14, 30, 14, 16, 24, 18, 50]); style_header(ws)

    # 7. LinkedIn ЛПР
    ws = wb.create_sheet('LinkedIn ЛПР')
    ws.append(['Компания', 'ИНН', 'В Bitrix?', 'ФИО', 'Должность', 'LinkedIn', 'Индустрия'])
    for li in linkedin:
        inn = str(li.get('inn') or '')
        ws.append([li.get('company', ''), inn, 'да' if inn in agg else 'нет', li.get('name', ''),
                   li.get('post', ''), li.get('linkedin', ''), li.get('industry', '')])
    widths(ws, [30, 14, 10, 24, 24, 40, 18]); style_header(ws)

    # 8. Глубокий рисёрч
    ws = wb.create_sheet('Глубокий рисёрч')
    ws.append(['Tier', 'Компания', 'ИНН', 'Досье (Perplexity)'])
    for inn, rec in research.items():
        ws.append([rec.get('tier', ''), rec.get('name', ''), inn, rec.get('text', '')[:2000]])
        ws.cell(ws.max_row, 4).alignment = WRAP
    widths(ws, [8, 30, 14, 120]); style_header(ws)

    # 9. Касания — лента
    ws = wb.create_sheet('Касания — лента')
    ws.append(['Дата', 'ИНН', 'Компания', 'Тип', 'Тема', 'Источник', 'Завершено?'])
    flat = []
    for inn, info in agg.items():
        for act in info['activities']:
            flat.append((act['date'] or '', inn, info['titles'][0] if info['titles'] else '', act['type'],
                         (act.get('subject') or '')[:80], act['source'], 'Да' if act['completed'] else 'Нет'))
    for r in sorted(flat, reverse=True):
        ws.append(list(r))
    widths(ws, [12, 14, 30, 12, 50, 20, 12]); style_header(ws)

    wb.save(a.out)
    print('Saved ->', a.out, '| sheets:', len(wb.sheetnames))

if __name__ == '__main__':
    main()
