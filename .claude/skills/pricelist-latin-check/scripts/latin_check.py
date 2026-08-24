#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Поиск латиницы в кириллическом тексте прайс-листов и каталогов (.xlsx).

Три категории находок:
  critical — латинская буква внутри кириллического слова (Система, образование).
             Визуально неотличимо, ломает поиск и сортировку. Всегда помечается.
  english  — целое латинское слово, которого нет в whitelist. Решает редакция.
  allowed  — бренд, аббревиатура, римская цифра. Не помечается, только считается.

Пишет копию файла с жёлтой заливкой проблемных ячеек и CSV со списком находок.
Исходный файл не изменяется.

Зависимости: openpyxl.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:  # pragma: no cover
    sys.exit("Нужен openpyxl:  pip install openpyxl")


CYR = r"Ѐ-ӿ"
LATIN_WORD = re.compile(r"[A-Za-z]+")
# латиница между кириллицей: образование
MIXED_INSIDE = re.compile(rf"[{CYR}][A-Za-z][{CYR}]")
# латиница в начале кириллического слова: Система
MIXED_START = re.compile(rf"(?:^|[\s\"«(\-/])([A-Za-z][{CYR}]+)")
# кириллическое слово, оканчивающееся латиницей: система -> ловится MIXED_INSIDE не всегда
MIXED_END = re.compile(rf"[{CYR}]+[A-Za-z](?![A-Za-z{CYR}])")
HAS_CYR = re.compile(rf"[{CYR}]")

WHITELIST = {
    # Римские цифры
    "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X",
    "XI", "XII", "XIII", "XIV", "XV", "XVI", "XVII", "XVIII", "XIX", "XX", "XXI",
    # Аббревиатуры
    "IT", "HR", "KPI", "OKR", "ESG", "SMM", "PR", "GR", "MBA", "CEO", "CTO", "CFO",
    "IPO", "SPO", "NFT", "BPM", "IQ", "EQ", "ORM", "BPMN", "FOREX", "STEAM",
    "SMART", "SPQR", "BaaS", "DeFi", "HBR", "B2B", "B2C", "AI", "ML", "UX", "UI",
    "SEO", "CRM", "ERP", "API", "PDF", "USB", "LED", "HD", "TV",
    # Торговые марки
    "ChatGPT", "OpenAI", "Toyota", "Starbucks", "Nike", "Amazon", "Nintendo",
    "Google", "Huawei", "Sony", "Apple", "Microsoft", "Excel", "Word", "Windows",
    "Agile", "Scrum", "Kanban", "Lean", "Nvidia", "Blackstone", "StoryBrand",
    "amoCRM", "Telegram", "YouTube", "LinkedIn",
    # Частые слова в названиях серий
    "Popular", "Science", "Harvard", "Business", "Review", "Guide",
    "Young", "Adult", "Top", "Fiction", "Non", "Pro", "Plus", "Mini", "Max",
    "New", "The", "of", "and", "for", "in", "on",
}

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def find_critical(text: str) -> list[str]:
    """Латинские буквы внутри кириллических слов."""
    hits = MIXED_INSIDE.findall(text)
    hits += MIXED_START.findall(text)
    hits += MIXED_END.findall(text)
    return hits


def classify(text: str, whitelist: set[str]) -> tuple[str, list[str]]:
    """-> ('critical' | 'english' | 'allowed' | 'clean', находки)."""
    if not LATIN_WORD.search(text):
        return "clean", []

    critical = find_critical(text)
    if critical:
        return "critical", sorted(set(critical))

    words = LATIN_WORD.findall(text)
    lower_wl = {w.lower() for w in whitelist}
    unknown = [w for w in words if w.lower() not in lower_wl]
    if unknown:
        return "english", sorted(set(unknown))
    return "allowed", sorted(set(words))


def guess_text_columns(ws, header_row: int, sample_rows: int = 200) -> list[int]:
    """Колонки, где больше половины непустых значений — строки длиннее 2 символов."""
    cols: list[int] = []
    last = min(ws.max_row, header_row + sample_rows)
    for col in range(1, ws.max_column + 1):
        filled = texty = 0
        for row in range(header_row + 1, last + 1):
            val = ws.cell(row=row, column=col).value
            if val is None or val == "":
                continue
            filled += 1
            if isinstance(val, str) and len(val.strip()) > 2:
                texty += 1
        if filled and texty / filled > 0.5:
            cols.append(col)
    return cols


def guess_header_row(ws, limit: int = 20) -> int:
    """Первая строка, где не меньше двух непустых строковых ячеек."""
    for row in range(1, min(ws.max_row, limit) + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        strings = [v for v in vals if isinstance(v, str) and v.strip()]
        if len(strings) >= 2:
            return row
    return 1


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file", help="путь к .xlsx")
    ap.add_argument("--sheet", help="имя листа (по умолчанию первый)")
    ap.add_argument("--columns", help="буквы колонок через запятую: E,F,S (иначе авто)")
    ap.add_argument("--whitelist", help="файл с допустимыми словами, по слову в строке")
    ap.add_argument("--only-critical", action="store_true", help="не помечать английские слова")
    ap.add_argument("--header-row", type=int, help="номер строки заголовков (иначе авто)")
    args = ap.parse_args()

    path = Path(args.file).expanduser()
    if not path.exists():
        print(f"Файл не найден: {path}")
        return 1

    whitelist = set(WHITELIST)
    if args.whitelist:
        wl_path = Path(args.whitelist).expanduser()
        extra = [ln.strip() for ln in wl_path.read_text(encoding="utf-8").splitlines()]
        whitelist |= {w for w in extra if w and not w.startswith("#")}

    wb = openpyxl.load_workbook(path)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    header_row = args.header_row or guess_header_row(ws)
    if args.columns:
        cols = [openpyxl.utils.column_index_from_string(c.strip()) for c in args.columns.split(",")]
    else:
        cols = guess_text_columns(ws, header_row)
    if not cols:
        print("Текстовых колонок не нашлось — укажи их вручную через --columns")
        return 1

    counts = {"critical": 0, "english": 0, "allowed": 0}
    findings: list[dict] = []

    for row in range(header_row + 1, ws.max_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not isinstance(val, str) or not val.strip():
                continue
            if not HAS_CYR.search(val):
                continue  # чисто латинская ячейка — это не смешение, пропускаем
            kind, hits = classify(val, whitelist)
            if kind == "clean":
                continue
            counts[kind] += 1
            if kind == "allowed":
                continue
            if kind == "english" and args.only_critical:
                continue
            cell.fill = YELLOW
            findings.append({
                "лист": ws.title,
                "ячейка": cell.coordinate,
                "категория": "опечатка" if kind == "critical" else "английское слово",
                "текст": val,
                "находки": " | ".join(hits),
            })

    out_xlsx = path.with_name(path.stem + "_ПРОВЕРКА" + path.suffix)
    out_csv = out_xlsx.with_suffix(".csv")
    wb.save(out_xlsx)
    with out_csv.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["лист", "ячейка", "категория", "текст", "находки"])
        writer.writeheader()
        writer.writerows(findings)

    letters = ", ".join(openpyxl.utils.get_column_letter(c) for c in cols)
    print(f"Лист: {ws.title} | колонки: {letters} | строк данных: {ws.max_row - header_row}")
    print(f"Критичные опечатки:  {counts['critical']:5d}")
    print(f"Английские слова:    {counts['english']:5d}")
    print(f"Допустимая латиница: {counts['allowed']:5d}  (не помечено)")
    print(f"Отчёт:  {out_xlsx}")
    print(f"Список: {out_csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
