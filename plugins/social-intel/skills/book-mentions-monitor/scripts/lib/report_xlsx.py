# -*- coding: utf-8 -*-
"""XLSX-отчёт в структуре Медиалогии. Листы: Сводные, Сообщения, СМИ по количеству/уровням,
Соцсети, Читательское, Регионы, Жанры, Слова (частотный словарь лемм), Каналы."""
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from lib.keyword_analyzer_ru import tokens as _tokens
except Exception:
    def _tokens(t): return re.findall(r"[а-яёa-z]+", (t or "").lower())

BRAND = "1F3864"
_hf = PatternFill("solid", fgColor=BRAND)
_wb = Font(bold=True, color="FFFFFF", size=10)
_thin = Side(style="thin", color="B0B0B0")
_bd = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
_lf = Alignment(horizontal="left", vertical="center", wrap_text=True)
_tit = Font(bold=True, size=13, color=BRAND)


def _sheet(wb, name, headers, rows, widths=None, title=None, left_cols=()):
    ws = wb.create_sheet(name[:31])
    r = 1
    if title:
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=max(len(headers), 1))
        ws.cell(1, 1, title).font = _tit
        r = 2
    for ci, h in enumerate(headers, 1):
        c = ws.cell(r, ci, h); c.fill = _hf; c.font = _wb; c.alignment = _ctr; c.border = _bd
        ws.column_dimensions[get_column_letter(ci)].width = (widths[ci - 1] if widths else 16)
    for row in rows:
        rr = ws.max_row + 1
        for ci, v in enumerate(row, 1):
            c = ws.cell(rr, ci, v); c.border = _bd
            c.alignment = _lf if ci in left_cols else _ctr
    ws.freeze_panes = ws.cell(r + 1, 1)
    return ws


def build(mentions, book, out_path, stopwords=None):
    stop = set(stopwords or [])
    rel = [m for m in mentions if m.get("_is_target")]
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    smi = [m for m in rel if m.get("_type") in ("СМИ", "Агрегатор")]
    soc = [m for m in rel if m.get("_type") == "Соцсеть"]
    rdr = [m for m in rel if m.get("_type") in ("Читательский", "Магазин")]
    vid = [m for m in rel if m.get("_type") == "Видео"]
    orig = [m for m in rel if not m.get("_reprint_of")]

    # 1. Сводные
    tone = Counter(m.get("_tone", "Нейтрал") for m in rel)
    _sheet(wb, "Сводные данные", ["Показатель", "Значение"], [
        ["Книга", book.get("title", "")],
        ["Авторы", ", ".join(book.get("authors", []))],
        ["Всего упоминаний (целевых)", len(rel)],
        ["Оригиналы / Перепечатки", f"{len(orig)} / {len(rel) - len(orig)}"],
        ["СМИ + агрегаторы", len(smi)],
        ["Соцсети", len(soc)],
        ["Видео (YouTube)", len(vid)],
        ["Читательское / Магазины", len(rdr)],
        ["Позитив", tone.get("Позитив", 0)],
        ["Нейтрал", tone.get("Нейтрал", 0)],
        ["Негатив", tone.get("Негатив", 0)],
        ["Индекс тональности", round((tone.get("Позитив", 0) - tone.get("Негатив", 0)) / max(len(rel), 1), 2)],
        ["Суммарный охват (оценка)", sum(m.get("_reach", 0) for m in rel)],
        ["Суммарный МедиаИндекс", round(sum(m.get("_mi", 0) for m in rel), 1)],
    ], widths=[32, 46], title="Сводные данные", left_cols=(1, 2))

    # 2. Сообщения
    rows = []
    for i, m in enumerate(sorted(rel, key=lambda x: x.get("date", ""), reverse=True), 1):
        rows.append([i, m.get("date", "")[:16], m.get("source", ""), m.get("_type", ""), m.get("_level", ""),
                     m.get("_category", ""), m.get("_city", ""), m.get("title", "")[:90], m.get("url", "")[:60],
                     m.get("_tone", ""), m.get("_role", ""), m.get("_genre", ""), m.get("_cite", ""),
                     "перепечатка" if m.get("_reprint_of") else "оригинал",
                     m.get("_reach", 0) or "н.д.", m.get("_mi", "")])
    _sheet(wb, "Сообщения", ["№", "Дата", "Источник", "Тип", "Уровень", "Категория", "Город", "Заголовок",
                              "Ссылка", "Тональность", "Роль", "Жанр", "Цит.", "Ориг/Переп", "Охват", "МедиаИндекс"],
           rows, widths=[5, 15, 22, 13, 13, 13, 13, 42, 30, 12, 13, 11, 7, 12, 12, 12],
           title="Сообщения (лента упоминаний)", left_cols=(3, 8, 9))

    # 3. СМИ по количеству
    src = Counter(m.get("source", "") or "—" for m in smi)
    _sheet(wb, "СМИ по количеству", ["Источник", "Кол-во", "Уровень"],
           [[s, c, next((m.get("_level", "") for m in smi if m.get("source") == s), "")] for s, c in src.most_common()],
           widths=[34, 8, 16], title="СМИ по количеству", left_cols=(1,))

    # 4. СМИ по уровням
    _sheet(wb, "СМИ по уровням", ["Уровень", "Кол-во"],
           list(Counter(m.get("_level", "") for m in smi).most_common()), widths=[20, 10], title="Упоминания по уровням СМИ")

    # 5. СМИ по категориям
    _sheet(wb, "СМИ по категориям", ["Категория", "Кол-во"],
           list(Counter(m.get("_category", "") for m in smi).most_common()), widths=[22, 10], title="Упоминания по категориям")

    # 6. Соцсети + Видео (с охватами)
    _sheet(wb, "Соцсети", ["Источник", "Канал", "Просмотры", "Лайки", "Репосты", "Ссылка"],
           [[m.get("source", ""), m.get("_category", ""), m.get("views") or "", m.get("likes") or "",
             m.get("reposts") or "", m.get("url", "")[:50]] for m in sorted(soc + vid, key=lambda x: (x.get("views") or 0), reverse=True)],
           widths=[24, 14, 11, 9, 9, 40], title="Соцсети и видео (с охватами)", left_cols=(1, 6))

    # 7. Читательское (рейтинги/отзывы)
    _sheet(wb, "Читательское", ["Площадка", "Рейтинг", "Оценок", "Заголовок/отзыв", "Ссылка"],
           [[m.get("source", "") or m.get("_category", ""), m.get("rating") or "", m.get("rating_count") or "",
             m.get("title", "")[:60], m.get("url", "")[:50]] for m in rdr],
           widths=[18, 9, 9, 50, 40], title="Читательский слой (LiveLib / магазины)", left_cols=(1, 4, 5))

    # 8. Регионы
    reg = Counter(m.get("_city", "") for m in rel if m.get("_city"))
    _sheet(wb, "Регионы", ["Город", "Кол-во"], list(reg.most_common()), widths=[24, 10], title="Упоминания по городам")

    # 9. Жанры
    _sheet(wb, "Жанры", ["Жанр", "Кол-во"], list(Counter(m.get("_genre", "") for m in rel).most_common()),
           widths=[20, 10], title="Упоминания по жанрам")

    # 10. Слова (частотный словарь лемм)
    words = Counter()
    for m in rel:
        for w in _tokens(m.get("title", "") + " " + m.get("snippet", "")):
            if w not in stop and len(w) > 3:
                words[w] += 1
    _sheet(wb, "Слова", ["Слово (лемма)", "Кол-во"], [[w, c] for w, c in words.most_common(40)],
           widths=[24, 10], title="Частотный словарь (лемматизация)")

    wb.save(out_path)
    return {"total": len(rel), "orig": len(orig), "smi": len(smi), "soc": len(soc), "rdr": len(rdr), "vid": len(vid),
            "sheets": wb.sheetnames, "path": str(out_path)}
